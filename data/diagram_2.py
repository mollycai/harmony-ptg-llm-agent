import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# ======================
# 1. 读取 Excel
# ======================
file_path = "data/results.xlsx"
wb = load_workbook(file_path)

# ======================
# 2. 百分比解析
# ======================
def parse_percent(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    else:
        return float(x) * 100 if float(x) <= 1 else float(x)

# ======================
# 3. 读取某个 sheet 的数据
# ======================
def load_sheet(sheet_name):
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[1:]

    times = np.array([r[0] for r in data])

    lp_sc = np.array([parse_percent(r[1]) for r in data])
    lp_pc = np.array([parse_percent(r[2]) for r in data])
    lp_an = np.array([r[3] for r in data])

    other_sc = np.array([parse_percent(r[4]) for r in data])
    other_pc = np.array([parse_percent(r[5]) for r in data])
    other_an = np.array([r[6] for r in data])

    return times, lp_sc, lp_pc, lp_an, other_sc, other_pc, other_an

# ======================
# 4. 配色（统一）
# ======================
COLOR_SC = "#FFC000"
COLOR_PC = "#FF0000"
COLOR_AN = "#4472C4"

plt.rcParams.update({
    "font.family": "serif",
    "font.size": 11,
    "axes.labelsize": 13,
    "legend.fontsize": 9,
    "figure.dpi": 200,
    "savefig.dpi": 400,
})

# ======================
# 5. 创建子图（关键）
# ======================
fig, axes = plt.subplots(1, 2, figsize=(12, 5))

sheet_names = ["LMT vs SAMT1", "LMT vs SAMT2"]
titles = ["(a) LMT vs SAMT1", "(b) LMT vs SAMT2"]

for i, (ax_left, sheet_name, title) in enumerate(zip(axes, sheet_names, titles)):

    # ---- 数据 ----
    times, lp_sc, lp_pc, lp_an, other_sc, other_pc, other_an = load_sheet(sheet_name)

    ax_right = ax_left.twinx()

    # ---- 左轴：AN ----
    l3, = ax_left.plot(times, lp_an, color=COLOR_AN, marker='o',
                      linestyle='-', label='LMT AN')

    l6, = ax_left.plot(times, other_an, color=COLOR_AN, marker='o',
                      linestyle='--', label='SAMT AN')

    # ---- 右轴：Coverage ----
    l1, = ax_right.plot(times, lp_sc, color=COLOR_SC, marker='x',
                       linestyle='-', label='LMT SC')

    l2, = ax_right.plot(times, lp_pc, color=COLOR_PC, marker='^',
                       linestyle='-', label='LMT PC')

    l4, = ax_right.plot(times, other_sc, color=COLOR_SC, marker='x',
                       linestyle='--', label='SAMT SC')

    l5, = ax_right.plot(times, other_pc, color=COLOR_PC, marker='^',
                       linestyle='--', label='SAMT PC')

    # ---- 轴标签 ----
    ax_left.set_xlabel("Time (sec.)")
    if i == 0:
        ax_left.set_ylabel("Action Number (AN)")
    if i == 1:
        ax_right.set_ylabel("Coverage (%)")

    # ---- 坐标范围（统一论文风格）----
    ax_right.set_ylim(20, 80)
    
    ax_right.set_ylim(
			min(lp_sc.min(), other_sc.min(), lp_pc.min(), other_pc.min()) - 2,
			max(lp_sc.max(), other_sc.max(), lp_pc.max(), other_pc.max()) + 2
		)

    # ---- 标题 ----
    ax_left.set_title(title)

    # ---- 网格 ----
    ax_left.grid(True, axis='y', linestyle='-', alpha=0.3)

    # ---- 图例（只放一张）----
    if i == 0:
        legend_lines = [l1, l2, l3, l4, l5, l6]
        legend_labels = [line.get_label() for line in legend_lines]

        ax_left.legend(
            legend_lines,
            legend_labels,
            loc='upper left',
            bbox_to_anchor=(0.02, 0.98),
            frameon=True
        )

# ======================
# 6. 布局 + 保存
# ======================
plt.tight_layout()

plt.savefig("data/diagram_2.png", bbox_inches="tight")

plt.show()