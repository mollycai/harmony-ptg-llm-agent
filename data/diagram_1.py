import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

# ======================
# 1. 读取 Excel
# ======================
file_path = "data/results.xlsx"   # 修改为你的路径
sheet_name = "LMT vs RT"

wb = load_workbook(file_path)
sheet = wb[sheet_name]

rows = list(sheet.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]

# ======================
# 2. 数据解析（自动处理 %）
# ======================
def parse_percent(x):
    if isinstance(x, str):
        return float(x.replace('%', ''))
    else:
        return float(x) * 100 if float(x) <= 1 else float(x)

times = np.array([r[0] for r in data])

lp_sc = np.array([parse_percent(r[1]) for r in data])
lp_pc = np.array([parse_percent(r[2]) for r in data])
lp_an = np.array([r[3] for r in data])

rand_sc = np.array([parse_percent(r[4]) for r in data])
rand_pc = np.array([parse_percent(r[5]) for r in data])
rand_an = np.array([r[6] for r in data])

# ======================
# 3. 配色（WPS风格）
# ======================
COLOR_SC = "#FFC000"   # 黄色
COLOR_PC = "#FF0000"   # 红色
COLOR_AN = "#4472C4"   # 蓝色

# ======================
# 4. 论文风格
# ======================
plt.rcParams.update({
    "font.family": "serif",
    "font.size": 12,
    "axes.labelsize": 16,
    "legend.fontsize": 11,
    "figure.dpi": 200,
    "savefig.dpi": 400,
})

# ======================
# 5. 双轴绘图
# ======================
fig, ax_left = plt.subplots(figsize=(8.2, 5.4))
ax_right = ax_left.twinx()

# ---- 左轴：AN ----
l3, = ax_left.plot(times, lp_an,
                  color=COLOR_AN, marker='o',
                  linestyle='-',
                  label='LLM PTG-based Testing AN')

l6, = ax_left.plot(times, rand_an,
                  color=COLOR_AN, marker='o',
                  linestyle='--',
                  label='Random Testing AN')

# ---- 右轴：Coverage ----
l1, = ax_right.plot(times, lp_sc,
                   color=COLOR_SC, marker='x',
                   linestyle='-',
                   label='LLM PTG-based Testing SC')

l2, = ax_right.plot(times, lp_pc,
                   color=COLOR_PC, marker='^',
                   linestyle='-',
                   label='LLM PTG-based Testing PC')

l4, = ax_right.plot(times, rand_sc,
                   color=COLOR_SC, marker='x',
                   linestyle='--',
                   label='Random Testing SC')

l5, = ax_right.plot(times, rand_pc,
                   color=COLOR_PC, marker='^',
                   linestyle='--',
                   label='Random Testing PC')

# ======================
# 6. 坐标轴（自适应）
# ======================
ax_left.set_xlabel("Time (sec.)")
ax_left.set_ylabel("Action Number (AN)")
ax_right.set_ylabel("Coverage (%)")

ax_right.set_ylim(20, 80)

ax_right.set_ylim(
    min(lp_sc.min(), rand_sc.min(), lp_pc.min(), rand_pc.min()) - 2,
    max(lp_sc.max(), rand_sc.max(), lp_pc.max(), rand_pc.max()) + 2
)

# 网格
ax_left.grid(True, axis='y', linestyle='-', alpha=0.3)

# ======================
# 7. 图例（关键点）
# ======================
legend_lines = [l1, l2, l3, l4, l5, l6]
legend_labels = [line.get_label() for line in legend_lines]

ax_left.legend(
    legend_lines,
    legend_labels,
    loc='upper left',
    bbox_to_anchor=(0.02, 0.98),   # 不遮挡关键
    frameon=True
)

# ======================
# 8. 保存
# ======================
plt.tight_layout()

plt.savefig("data/diagram_1.png", bbox_inches="tight")  # 预览

plt.show()